import difflib
import json

import openpyxl


#将short response转换成数字，并判断答案是否正确
#将short response转换成对应的数字，从1开始（manual label更好标记，不易出错）
def short_response2num(options,short_response):#options:问题选项；short_response:得到的回答；9表示不知道，0表示给出了不是选项里的答案
    options.append("i'm sorry, don't know")
    short_response = short_response.lower().strip()
    # 存储匹配分数
    highest_similarity = 0.0
    option_num=0
    for i, option in zip(range(0,len(options)), options):
        similarity = difflib.SequenceMatcher(None, short_response, option.lower()).ratio()
        # 更新最高相似度和最佳匹配
        if similarity > highest_similarity:
            highest_similarity = similarity
            option_num=i+1
    if option_num==len(options):
        option_num=9
    if highest_similarity>=0.5:
        return option_num
    else:
        return 0

#将正确答案转换为对应的数字，从1开始
def correct_answer2num(options,correct_answer):#options:问题选项；correct_answer:正确答案
    correct_answer=correct_answer.lower().strip()
    highest_similarity=0
    answer_num=0
    for i, option in zip(range(0,len(options)), options):
        similarity = difflib.SequenceMatcher(None, correct_answer, option.lower()).ratio()
        # 更新最高相似度和最佳匹配
        if similarity > highest_similarity:
            highest_similarity = similarity
            answer_num=i+1
    return answer_num

#从json文件中得到对应的选项，正确答案和回复,并将处理后的结果写入excel
def NEJM_result2excel(initial_json,output_json):
    with open(initial_json, 'r', encoding='utf-8') as file:
        initial_json_data=json.load(file)
    with open(output_json, 'r', encoding='utf-8') as file:
        output_json_data=json.load(file)
    NEJM_excel_sheet=[]
    cases_options = []
    i=1
    for initial_case_data,output_case_data in zip(initial_json_data,output_json_data):
        #options保存每个case的选项（文本）
        options=[]
        options_data=initial_case_data["options"]
        for option in options_data:
            options.append(option["option"])
        correct_answer=initial_case_data["correct answer"]
        short_response = output_case_data["Short_response"]
        initial_case_id=initial_case_data["image name"]
        output_case_id=output_case_data["image_name"]
        if initial_case_id==output_case_id:
            correct_answer_num = correct_answer2num(options, correct_answer)
            option_num=short_response2num(options,short_response)
            NEJM_excel_sheet_data=[i,output_case_id,correct_answer_num,option_num,correct_answer]
            NEJM_excel_sheet.append(NEJM_excel_sheet_data)
            #NEJM_excel_sheet.append(json.dumps(NEJM_excel_sheet_data, ensure_ascii=False))
            cases_options_data=options
            cases_options.append(cases_options_data)
            #cases_options.append(json.dumps(cases_options_data, ensure_ascii=False))
        i+=1
    return NEJM_excel_sheet,cases_options

def Lancet_result2excel(initial_json,output_json):
    with open(initial_json, 'r', encoding='utf-8') as file:
        initial_json_data = json.load(file)
    with open(output_json, 'r', encoding='utf-8') as file:
        output_json_data = json.load(file)
    Lancet_excel_sheet = []
    cases_options = []
    i = 1
    for initial_case_data, output_case_data in zip(initial_json_data, output_json_data):
        # options保存每个case的选项（文本）
        options = []
        options_data = initial_case_data["options"]
        for option in options_data:
            options.append(option)
        # 保存正确答案的文本
        correct_answer = initial_case_data["correct answer"]
        short_response = output_case_data["Short_response"]
        initial_case_id = initial_case_data["image"][0]["image name"]
        output_case_id = output_case_data["image_name"][0]
        if initial_case_id == output_case_id:
            correct_answer_num = correct_answer2num(options, correct_answer)
            option_num = short_response2num(options, short_response)
            Lancet_excel_sheet_data = [i, output_case_id, correct_answer_num, option_num, correct_answer]
            Lancet_excel_sheet.append(Lancet_excel_sheet_data)
            cases_options_data = options
            cases_options.append(cases_options_data)
        i += 1
    return Lancet_excel_sheet, cases_options

def JAMA_result2excel(initial_json,output_json):
    with open(initial_json, 'r', encoding='utf-8') as file:
        initial_json_data=json.load(file)
    with open(output_json, 'r', encoding='utf-8') as file:
        output_json_data=json.load(file)
    JAMA_excel_sheet=[]
    cases_options = []
    i=1
    for initial_case_data,output_case_data in zip(initial_json_data,output_json_data):
        options=[]
        options_data=initial_case_data["options"]
        for option in options_data:
            options.append(option)
        #保存正确答案的文本
        correct_answer=initial_case_data["correct answer"]
        short_response = output_case_data["Short_response"]
        initial_case_id=initial_case_data["image name"]
        output_case_id=output_case_data["image_name"]
        if initial_case_id==output_case_id:
            # 正确答案转为数字
            correct_answer_num = correct_answer2num(options, correct_answer)
            option_num=short_response2num(options,short_response)
            JAMA_excel_sheet_data=[i,output_case_id,correct_answer_num,option_num,correct_answer]
            JAMA_excel_sheet.append(JAMA_excel_sheet_data)
            cases_options_data=options
            cases_options.append(cases_options_data)
        i+=1
    return JAMA_excel_sheet,cases_options


######################################
#                                    #
#               NEJM                 #
#                                    #
######################################
'''
workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "NEJM_Label_tem0"#修改1
sheet1.append(["ID","image_name","correct_answer_num","option_num","correct_answer"])
initial_json="textdata/NEJM.json"
output_json="output/NEJM/role-play/NEJM_output_tem0.json"#修改2
NEJM_excel_sheet,cases_options=NEJM_result2excel(initial_json,output_json)
for NEJM_excel_sheet_data in NEJM_excel_sheet:
    sheet1.append(NEJM_excel_sheet_data)

workbook.save("auto_label/NEJM/role-play/NEJM_Label_tem0.xlsx")#修改3
sheet2 = workbook.active
sheet2.title = "NEJM_options_tem0"#修改4
for case_options in cases_options:
    sheet2.append(case_options)
workbook.save("auto_label/NEJM/role-play/NEJM_Label_tem0.xlsx")#修改5
'''

######################################
#                                    #
#               Lancet               #
#                                    #
######################################
'''
workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "Lancet_Label_tem0"#修改1
sheet1.append(["ID","image_name","correct_answer_num","option_num","correct_answer"])
initial_json="textdata/Lancet.json"
output_json="output/Lancet/none role-play/Lancet_output_tem0.json"#修改2
Lancet_excel_sheet,cases_options=Lancet_result2excel(initial_json,output_json)
for Lancet_excel_sheet_data in Lancet_excel_sheet:
    sheet1.append(Lancet_excel_sheet_data)
    #修改excel的保存路径
workbook.save("auto_label/Lancet/none role-play/Lancet_Label_tem0.xlsx")#修改3
sheet2 = workbook.active
sheet2.title = "Lancet_options_tem0"#修改4
for case_options in cases_options:
    sheet2.append(case_options)
workbook.save("auto_label/Lancet/none role-play/Lancet_Label_tem0.xlsx")#修改5
'''


######################################
#                                    #
#               JAMA                 #
#                                    #
######################################
workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "JAMA_Label_tem0.5"#修改1
sheet1.append(["ID","image_name","correct_answer_num","option_num","correct_answer"])
initial_json="textdata/JAMA.json"
output_json="output/JAMA/role-play/JAMA_output_tem0.5.json"#修改2
JAMA_excel_sheet,cases_options=JAMA_result2excel(initial_json,output_json)
for JAMA_excel_sheet_data in JAMA_excel_sheet:
    sheet1.append(JAMA_excel_sheet_data)
workbook.save("auto_label/JAMA/role-play/JAMA_Label_tem0.5.xlsx")#修改3
sheet2 = workbook.active
sheet2.title = "JAMA_options_tem0.5"#修改4
for case_options in cases_options:
    sheet2.append(case_options)
workbook.save("pretest/auto_label/JAMA/role-play/JAMA_Label_tem0.5.xlsx")#修改5